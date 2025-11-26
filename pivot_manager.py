import logging
import shutil
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook

# Configuração de Logs para Rastreabilidade
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - [%(class)s] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger("PivotManager")

class PivotTableConfigurator:
    """
    Gerenciador de Tabelas Dinâmicas via manipulação direta de arquivos OpenXML (.xlsx).
    Foca em configurar metadados para comportamento automático no Excel.
    """

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.extra_log_info = {'class': self.__class__.__name__}

    def _create_backup(self) -> None:
        """Cria backup de segurança antes de manipular o XML do arquivo."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.file_path}")
            
        backup_path = self.file_path.with_suffix(f".backup{self.file_path.suffix}")
        shutil.copy2(self.file_path, backup_path)
        logger.info(f"Backup de segurança criado em: {backup_path}", extra=self.extra_log_info)

    def set_refresh_on_load(self, sheet_name: str, pivot_name: Optional[str] = None) -> bool:
        """
        Configura a Tabela Dinâmica para atualizar automaticamente ao abrir o Excel.
        
        Args:
            sheet_name (str): Nome da aba onde está a tabela.
            pivot_name (str, opcional): Nome específico da tabela. Se None, aplica em todas da aba.
        """
        try:
            self._create_backup()
            
            logger.info(f"Carregando estrutura do arquivo: {self.file_path.name}...", extra=self.extra_log_info)
            # data_only=False é crucial para manter fórmulas
            wb = load_workbook(self.file_path, keep_vba=True)
            
            if sheet_name not in wb.sheetnames:
                logger.error(f"Aba '{sheet_name}' não encontrada.", extra=self.extra_log_info)
                return False

            ws = wb[sheet_name]
            pivot_found = False

            # As tabelas dinâmicas no openpyxl são acessadas via pivot_tables
            # Nota: Dependendo da versão do openpyxl, a iteração pode variar.
            # Esta abordagem foca na manipulação da definição de cache.
            
            for pivot in ws._pivots:
                if pivot_name and pivot.name != pivot_name:
                    continue
                
                # A Mágica: Acessamos a definição de cache e forçamos o refresh
                pivot.cache.refreshOnLoad = True
                pivot_found = True
                logger.info(f"Configurado 'RefreshOnLoad=True' para: {pivot.name}", extra=self.extra_log_info)

            if not pivot_found:
                logger.warning(f"Nenhuma tabela dinâmica encontrada ou correspondente na aba {sheet_name}.", extra=self.extra_log_info)
                return False

            logger.info("Salvando alterações no arquivo...", extra=self.extra_log_info)
            wb.save(self.file_path)
            logger.info("✅ Arquivo atualizado com sucesso.", extra=self.extra_log_info)
            return True

        except Exception as e:
            logger.exception(f"Falha crítica ao processar planilha: {e}", extra=self.extra_log_info)
            return False
        finally:
            # openpyxl não exige 'close' explícito como win32, mas é bom liberar memória se o objeto for grande
            if 'wb' in locals():
                wb.close()

if __name__ == "__main__":
    # Configuração de Entrada
    ARQUIVO_ALVO = r"C:\Dados\Relatorio_Vendas.xlsx"
    NOME_ABA = "Geral"
    NOME_TABELA = "TabelaDinamica1" # Use None para atualizar todas na aba

    configurator = PivotTableConfigurator(ARQUIVO_ALVO)
    configurator.set_refresh_on_load(NOME_ABA, NOME_TABELA)
